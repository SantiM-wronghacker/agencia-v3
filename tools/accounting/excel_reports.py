from pathlib import Path

from tools.base import BaseTool, ToolResult, tool


@tool
class ExcelReportsTool(BaseTool):
    name = "excel_reports"
    description = (
        "Genera y lee reportes en Excel. "
        "Úsala para crear reportes financieros, "
        "de ventas o cualquier dato tabular."
    )

    def run(self, action: str = "", **kwargs) -> ToolResult:
        if action == "generate":
            return self._generate(**kwargs)
        if action == "read":
            return self._read(**kwargs)
        return self._error(f"Acción '{action}' no soportada")

    def _generate(
        self,
        title: str = "",
        headers: list = None,
        rows: list = None,
        output_path: str = None,
    ) -> ToolResult:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return self._error(
                "openpyxl no instalado. Ejecuta: pip install openpyxl"
            )

        headers = headers or []
        rows = rows or []

        if output_path is None:
            safe_title = title[:30].replace(" ", "_")
            output_path = f"data/docs/{safe_title}.xlsx"

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit is 31 chars

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        try:
            wb.save(output_path)
        except Exception as e:
            return self._error(f"Error al guardar Excel: {e}")

        return self._success(
            f"Excel generado: {output_path} ({len(rows)} filas)",
            raw_data={"path": output_path, "rows": len(rows)},
        )

    def _read(self, file_path: str = "", sheet_name: str = None) -> ToolResult:
        try:
            import openpyxl
        except ImportError:
            return self._error(
                "openpyxl no instalado. Ejecuta: pip install openpyxl"
            )

        if not file_path or not Path(file_path).exists():
            return self._error(f"Archivo no encontrado: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
        except Exception as e:
            return self._error(f"Error al abrir Excel: {e}")

        rows_data = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows_data.append([str(c) if c is not None else "" for c in row])

        if not rows_data:
            return self._success("Hoja vacía")

        # Format as text table
        lines = []
        for row in rows_data:
            lines.append(" | ".join(row))

        return self._success(
            "\n".join(lines),
            raw_data={"rows": rows_data, "total": len(rows_data)},
        )
